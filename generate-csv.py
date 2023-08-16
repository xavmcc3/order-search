from dotenv import load_dotenv
from clrprint import clrprint
from pathlib import Path
import openpyxl
import csv
import os

def write_excel_to_csv(excel_path, csv_path):
    clrprint("Serializing", excel_path, end="", clr="w,m")
    print("...")

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    with open(csv_path, 'a', newline="") as f:
        c = csv.writer(f)
        for row in ws.iter_rows(min_row=2, max_col=11):
            if row[0].value == None and row[1].value == None:
                continue
            values = []
            for cell in row:
                values.append(cell.value)
            c.writerow(values)
        f.truncate()
    wb.close()

if __name__ == "__main__":
    load_dotenv()
    csv_path = "./data/index.csv"
    folder = os.getenv("EXCEL_DIR")

    open(csv_path, 'w').close()
    for file in os.listdir(folder):
        if not file.endswith(".xlsx"):
            continue
        if file.startswith("~$"):
            continue
        path = os.path.abspath(os.path.join(folder, file))
        try:
            write_excel_to_csv(path, csv_path)
        except Exception as e:
            clrprint("[ERROR]", "in", path, clr="r,w,m")
            print(e)
    clrprint("[DONE]", "Operator spreadsheets serialized.", clr="g,w")