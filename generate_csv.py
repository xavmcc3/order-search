from dotenv import load_dotenv
from datetime import datetime
from clrprint import clrprint
from pathlib import Path
import openpyxl
import csv
import os

def write_excel_to_csv(excel_path, csv_path, check_blanks=True, max_blanks=1):
    clrprint("Serializing", excel_path, end="", clr="w,m")
    print("...")

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    with open(csv_path, 'a', newline="") as f:
        c = csv.writer(f)
        blanks = 0
        for row in ws.iter_rows(min_row=2, max_col=11):
            if row[0].value == None and row[1].value == None:
                if check_blanks:
                    continue
                blanks += 1
                if blanks > max_blanks:
                    break
                continue
            values = []
            for cell in row:
                values.append(cell.value)
            c.writerow(values)
        f.truncate()
    wb.close()
    return f"> Serialized <color-m>{excel_path}<color-m>".replace("\\", '/')

async def generate_csv(folder_path, csv_path, output=lambda _ : "", check_blanks=True, max_blanks=1):
    folder = folder_path

    start_time = datetime.now()
    open(csv_path, 'w').close()
    for file in os.listdir(folder):
        if not file.endswith(".xlsx"):
            continue
        if file.startswith("~$"):
            continue
        path = os.path.abspath(os.path.join(folder, file))
        try:
            output(write_excel_to_csv(path, csv_path, check_blanks=check_blanks, max_blanks=max_blanks))
        except Exception as e:
            clrprint("[ERROR]", "in", path, clr="r,w,m")
            print(e)

            output(f"[ERROR] in {path}".replace("\\", '/'), True, False)
            output(str(e), True)
            output(f"<color-r>[ERROR]</color-r> in <color-m>{path}</color-m>".replace("\\", '/'))
    clrprint("[DONE] ", "Operator spreadsheets serialized in ", f"{datetime.now() - start_time}", ".", sep="", clr="g,w,y,w")

if __name__ == "__main__":
    load_dotenv()
    folder = os.getenv("EXCEL_DIR")
    generate_csv(folder)
